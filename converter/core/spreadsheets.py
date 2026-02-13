"""Excel (XLSX) <-> CSV conversions."""

import csv
from pathlib import Path

from converter.utils.helpers import get_output_path, validate_input_file


def xlsx_to_csv(input_path: str, output_dir: str | None = None, sheet_name: str | None = None) -> Path:
    """Convert an Excel file to CSV.

    If sheet_name is not specified, the active (first) sheet is used.
    """
    from openpyxl import load_workbook

    input_file = validate_input_file(input_path)
    output_path = get_output_path(input_path, "csv", output_dir)

    wb = load_workbook(str(input_file), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    wb.close()
    return output_path


def csv_to_xlsx(input_path: str, output_dir: str | None = None) -> Path:
    """Convert a CSV file to Excel (XLSX)."""
    from openpyxl import Workbook

    input_file = validate_input_file(input_path)
    output_path = get_output_path(input_path, "xlsx", output_dir)

    wb = Workbook()
    ws = wb.active

    with open(input_file, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save(str(output_path))
    return output_path
